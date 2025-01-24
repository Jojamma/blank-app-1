import streamlit as st
import pandas as pd
import os
from openpyxl import Workbook, load_workbook

# Title of the app
st.title("Large File Uploader with Excel Logging")

# File uploader for dataset
uploaded_file = st.file_uploader("Upload your dataset (supports large files up to 50GB)", type=None)

# Dropdown for model type selection
model_type = st.selectbox("Select Model Type:", ["Transformer", "CNN", "RNN"])

# Dropdown for core options selection
core_option = st.selectbox("Select Core Option:", ["CPU", "GPU", "HDFS"])

# Button to process the uploaded file
if st.button("Run"):
    if uploaded_file is not None:
        # Check if the uploaded file is a CSV
        if uploaded_file.name.endswith('.csv'):
            try:
                # Read and process CSV in chunks to avoid memory issues
                st.write("Processing CSV file...")
                chunk_size = 10 ** 6  # Adjust chunk size as needed
                total_rows = 0
                
                for chunk in pd.read_csv(uploaded_file, chunksize=chunk_size):
                    total_rows += len(chunk)
                
                # Display column names after processing all chunks
                df = pd.read_csv(uploaded_file)
                column_names = df.columns.tolist()
                st.write("CSV Column Names:", column_names)
                
            except Exception as e:
                st.error(f"Error reading CSV file: {e}")
        
        # Check if it's an image or other file types
        elif uploaded_file.name.endswith(('.jpg', '.jpeg', '.png')):
            st.write("Image Data: ", uploaded_file.name)
        
        else:
            st.write("File uploaded successfully but not recognized as a CSV or image.")

        # Get dataset size in MB
        dataset_size_mb = round(os.path.getbuffer(uploaded_file).nbytes / (1024 * 1024), 2)

        # Display selected model type and core option
        st.write(f"Model Type: {model_type}")
        st.write(f"Core Option: {core_option}")

        # Create or update an Excel file with dataset size and model name
        excel_file = "log.xlsx"
        if not os.path.exists(excel_file):
            # Create a new Excel file if it doesn't exist
            wb = Workbook()
            ws = wb.active
            ws.title = "Log"
            ws.append(["Date", "Dataset Size (MB)", "Model Name"])
            wb.save(excel_file)

        # Update the Excel file with new data
        wb = load_workbook(excel_file)
        ws = wb["Log"]
        ws.append([pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S"), dataset_size_mb, model_type])
        wb.save(excel_file)

        # Provide a download link for the Excel file
        with open(excel_file, "rb") as f:
            st.download_button(
                label="Download Log Excel File",
                data=f,
                file_name="log.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    else:
        st.error("Please upload a valid file.")
